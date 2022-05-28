print('Initializing...')
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from tqdm import tqdm as tq     #This package allows us to track loop progress
import time
import pandas as pd
import numpy as np
from functools import reduce

workbook0 = input('Name of data file to convert (please include .xlsx): ')
workbookf = input('Save converted data as (please include .xlsx): ')
print('Downloading Workbook...')

wb = load_workbook(workbook0)  #load workbook
ws = wb.active  #Assign active sheet as ws

#step1)change names of demographic categories
#creating separate functions for each column, each function makes the necessary changes

def regroup_age(r):     #renames age groups (frequency descending to save processing time)
    age_cell = ws.cell(row=r, column=3).value
    if age_cell == '18 to 49 years':
        ws.cell(row=r, column=3).value = 'c1849'
    elif age_cell == '50 to 64 years':
        ws.cell(row=r, column=3).value = 'c5064'
    elif age_cell == '0 - 17 years':
        ws.cell(row=r, column=3).value = 'c017'
    elif age_cell == '65+ years':
        ws.cell(row=r, column=3).value = 'c65up'
    else:
        ws.cell(row=r, column=3).value = 'cUkA'

def regroup_sex(r):     #renames sex groups
    sex_cell = ws.cell(row=r, column=4).value
    if sex_cell == 'Female':
        ws.cell(row=r, column=4).value = 'cF'
    elif sex_cell == 'Male':
        ws.cell(row=r, column=4).value = 'cM'
    else:
        ws.cell(row=r, column=4).value = 'cUkS'

def regroup_race(r):    #renames race groups
    race_cell = ws.cell(row=r, column=5).value
    if race_cell == 'NA': 
        ws.cell(row=r, column=5).value = 'cUkR'   
    elif race_cell == 'White':
        ws.cell(row=r, column=5).value = 'cW'
    elif race_cell == 'American Indian/Alaska Native':
        ws.cell(row=r, column=5).value = 'cNA'
    elif race_cell == 'Unknown' or ws.cell(row=r, column=5).value == 'Missing':
        ws.cell(row=r, column=5).value = 'cUkR'
    elif race_cell == 'Multiple/Other':
        ws.cell(row=r, column=5).value = 'cOth'
    elif race_cell == 'Black':
        ws.cell(row=r, column=5).value = 'cB'
    else:
        ws.cell(row=r, column=5).value = 'cA'

def regroup_ethnicity(r):   #renames ethnicity groups
    eth_cell = ws.cell(row=r, column=6).value
    if eth_cell == 'NA':
       ws.cell(row=r, column=6).value = 'cUkH'
    elif eth_cell == 'Non-Hispanic/Latino':
        ws.cell(row=r, column=6).value = 'cNH'
    elif eth_cell == 'Hispanic/Latino':
        ws.cell(row=r, column=6).value = 'cH'
    else:
        ws.cell(row=r, column=6).value = 'cUkH'
        

rows = ws.max_row       #number of occupied rows in spreadsheet (includes header)

for r in tq(range(2,rows+1), desc= "Age Regrouping"):  #iterate from row 2 to last occupied row, include progress bar
    regroup_age(r)
print('Age Regrouping Complete')
time.sleep(1)

for r in tq(range(2,rows+1), desc= "Sex Regrouping"):  #iterate from row 2 to last occupied row, include progress bar
    regroup_sex(r)
print('Sex Regrouping Complete')
time.sleep(1) 

for r in tq(range(2,rows//2), desc= "Race Regrouping 1/2"):     #Splitting race regrouping in 2 loops (program slows down here when all 1 loop)
    regroup_race(r)
print('Race Regrouping 1/2 Complete')
time.sleep(1)

for r in tq(range(rows//2,rows+1), desc= "Race Regrouping 2/2"):
    regroup_race(r)
print('Race Regrouping Complete')
time.sleep(1)
      
for r in tq(range(2,rows+1), desc= "Ethnicity Regrouping"):
    regroup_ethnicity(r)
print('Ethnicity Regrouping Complete')
time.sleep(1)

#step2)Change names of counties to all caps, change 'DOÃ‘A ANA’ to DONA ANA
#using same method as above:

def rename_county(r):   #Capitalizes county names, fixes 'DOÃ‘A ANA’ county
    county = ws.cell(row=r, column=1).value
    if county[0] == 'D': ws.cell(row=r, column=1).value = 'DONA ANA'
    county.upper()

for r in tq(range(2,rows+1), desc= "County Renaming"):
    rename_county(r)
print('County Rename Complete')
time.sleep(1)

#Step3) Add POC column, values either 1 or 0

ws.insert_cols(7)   #Insert new column
ws.cell(row=1, column=7).value = 'POC'  #create col. header

def flag_colPOC(r):
    POC_cell = ws.cell(row=r, column=7).value 
    eth_cell = ws.cell(row=r, column=6).value
    race_cell = ws.cell(row=r, column=5).value
    POC_race = ['cNA', 'cB', 'cA', 'cOth']      #array of race categories which are flagged as POC
    if eth_cell == 'cH' or race_cell in POC_race:
        ws.cell(row=r, column=7).value = 1
    else:
        ws.cell(row=r, column=7).value = 0

for r in tq(range(2,rows+1), desc= "Creating POC Column"):
    flag_colPOC(r)
print('POC Column Complete')
time.sleep(1)

#step4) Create flag columns for derived rece/ethnic groups in same manner as POC

ws.insert_cols(8,9)
ws.insert_cols(10,11)
ws.cell(row=1, column=8).value = 'cWNH'
ws.cell(row=1, column=9).value = 'cUkHAR'
ws.cell(row=1, column=10).value = 'cUkRH'
ws.cell(row=1, column=11).value = 'cWH'

def flag_wnh(r):
    wnh_cell = ws.cell(row=r, column=8).value 
    eth_cell = ws.cell(row=r, column=6).value
    race_cell = ws.cell(row=r, column=5).value
    if eth_cell == 'cNH' and race_cell == 'cW' :
        ws.cell(row=r, column=8).value = 1
    else:
        ws.cell(row=r, column=8).value = 0

def flag_unkhar(r):
    unkhar_cell = ws.cell(row=r, column=9).value 
    eth_cell = ws.cell(row=r, column=6).value
    if eth_cell == 'cUkH':
        ws.cell(row=r, column=9).value = 1
    else:
        ws.cell(row=r, column=9).value = 0

def flag_ukrh(r):
    ukrh_cell = ws.cell(row=r, column=10).value
    race_cell = ws.cell(row=r, column=5).value
    eth_cell = ws.cell(row=r, column=6).value
    if eth_cell == 'cUkH' and race_cell == 'cUkR':
        ws.cell(row=r, column=10).value = 1
    else:
        ws.cell(row=r, column=10).value = 0

def flag_wh(r):
    wh_cell = ws.cell(row=r, column=11).value
    race_cell = ws.cell(row=r, column=5).value
    eth_cell = ws.cell(row=r, column=6).value
    if eth_cell == 'cH' and race_cell == 'cW':
        ws.cell(row=r, column=11).value = 1
    else:
        ws.cell(row=r, column=11).value = 0

#These loops are all split in 2 for same reason as race regrouping loop
for r in tq(range(2,rows//2), desc= "Creating Derived Column 1"):
    flag_wnh(r)
print('Derived Column 1: 1/2 Complete')
time.sleep(1)

for r in tq(range(rows//2,rows+1), desc= "Creating Derived Column 1"):
    flag_wnh(r)
print('Derived Column 1: 2/2 Complete')
time.sleep(1)

for r in tq(range(2,rows//2), desc= "Creating Derived Column 2"):
    flag_unkhar(r)
print('Derived Column 2: 1/2 Complete')
time.sleep(1)

for r in tq(range(rows//2,rows+1), desc= "Creating Derived Column 2"):
    flag_unkhar(r)
print('Derived Column 2: 2/2 Complete')
time.sleep(1)

for r in tq(range(2,rows//2), desc= "Creating Derived Column 3"):
    flag_ukrh(r)
print('Derived Column 3: 1/2 Complete')
time.sleep(1)

for r in tq(range(rows//2,rows+1), desc= "Creating Derived Column 3"):
    flag_ukrh(r)
print('Derived Column 3: 2/2 Complete')
time.sleep(1)

for r in tq(range(2,rows//2), desc= "Creating Derived Column 4"):
    flag_wh(r)
print('Derived Column 4: 1/2 Complete')
time.sleep(1)

for r in tq(range(rows//2,rows+1), desc= "Creating Derived Column 4"):
    flag_wh(r)
print('Derived Column 4: 2/2 Complete')

#print('Saving as COVIDdemofile_presum.xlsx ...')
#wb.save('COVIDdemofile_presum.xlsx')    #saving the altered data in new file before summarizing
print('Phase 1 complete!')

#step4b) Summarize table 5 times. Once for each of columns age_group, sex, race, ethnicity, POC
#print('Downloading Case Counts...')
#md = pd.read_excel('COVIDdemofile_presum.xlsx')
print('Converting to summarizable dataframe...')
data = ws.values
columns = next(data)[0:]
md = pd.DataFrame(data, columns=columns)    #This converts openpyxl ws into pandas df

print('Downloading Population Counts...')
pc = pd.read_excel('PopulationCounts.xlsx')

print ('Summarizing All Demographics...')
summ_age = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='age_group', aggfunc='size', fill_value=0)

summ_sex = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='sex', aggfunc='size', fill_value=0)

summ_race = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='race', aggfunc='size', fill_value=0)

summ_eth = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='ethnicity', aggfunc='size', fill_value=0)

summ_POC = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='POC', aggfunc='size', fill_value=0)
summ_POC.drop(columns=0, inplace=True)
summ_POC.rename(columns={1: 'cPOC'}, inplace=True)

summ_WNH = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='cWNH', aggfunc='size', fill_value=0)
summ_WNH.drop(columns=0, inplace=True)
summ_WNH.rename(columns={1: 'cWNH'}, inplace=True)

summ_UkHAR = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='cUkHAR', aggfunc='size', fill_value=0)
summ_UkHAR.drop(columns=0, inplace=True)
summ_UkHAR.rename(columns={1: 'cUkHAR'}, inplace=True)

summ_UkRH = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='cUkRH', aggfunc='size', fill_value=0)
summ_UkRH.drop(columns=0, inplace=True)
summ_UkRH.rename(columns={1: 'cUkRH'}, inplace=True)

summ_WH = pd.pivot_table(md, index=['res_county', 'county_fips_code'], columns='cWH', aggfunc='size', fill_value=0)
summ_WH.drop(columns=0, inplace=True)
summ_WH.rename(columns={1: 'cWH'}, inplace=True)


print('Merging Demographic Tables...')
summs = [summ_age, summ_sex, summ_race, summ_eth, summ_POC, summ_WNH, summ_UkHAR, summ_UkRH, summ_WH]
summ_merge = reduce(lambda  left,right: pd.merge(left,right,on=['res_county', 'county_fips_code']), summs)

dfm = pd.merge(summ_merge,pc, how='inner', on='res_county')

print('Calculating Rates and Percentages...')

def rate(c,pop,rate):
    dfm[rate] = np.where(pop !=0,(c/pop)*100, 'NA')

def percent(c,ctot,per):
    dfm[per] = (c/ctot)*100

dfm['cTot'] = dfm.cM + dfm.cF + dfm.cUkS
#dfm['popUkA'] = dfm.popTotal - (dfm.pop017 + dfm.pop1849 + dfm.pop5064 + dfm.pop65up)
#dfm['popUkS'] = dfm.popTotal - (dfm.popM + dfm.popF)


cs = [dfm.c017, dfm.c1849,dfm.c5064,dfm.c65up,dfm.cUkA,dfm.cF,dfm.cM,dfm.cUkS,dfm.cA,dfm.cB,dfm.cNA,dfm.cOth,dfm.cUkR
         ,dfm.cW,dfm.cH,dfm.cNH,dfm.cUkH,dfm.cPOC,dfm.cWNH,dfm.cUkHAR,dfm.cUkRH,dfm.cWH]
pops = [dfm.pop017, dfm.pop1849,dfm.pop5064,dfm.pop65up,0,dfm.popF,dfm.popM,0,dfm.popA,dfm.popB,dfm.popNA,0,0
         ,0,dfm.popH,0,0,0,dfm.popWNH,0,0,dfm.popWH]
rates = ['r017', 'r1849','r5064','r65up','rUkA','rF','rM','rUkS','rA','rB','rNA','rOth','rUkR'
         ,'rW','rH','rNH','rUkH','rPOC','rWNH','rUkHAR','rUkRH','rWH']
pers = ['p017', 'p1849','p5064','p65up','pUkA','pF','pM','pUkS','pA','pB','pNA','pOth','pUkR'
         ,'pW','pH','pNH','pUkH','pPOC','pWNH','pUkHAR','pUkRH','pWH']
ctot = dfm.cTot

for i in range(len(cs)):
    rate(cs[i],pops[i],rates[i])
    percent(cs[i],ctot,pers[i])

dfm.to_excel(workbookf)
print('Complete.')

time.sleep(5)





        
    




    
    
    
        


